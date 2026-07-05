import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from table_ocr.models import TableRow


HEADERS = ["source_image", "unit_no", "donor_name", "blood_group", "contact_no", "confidence", "needs_review", "raw_ocr"]


def export_rows(rows: list[TableRow], output_stem: Path) -> tuple[Path, Path]:
    output_stem.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_stem.with_suffix(".csv")
    xlsx_path = output_stem.with_suffix(".xlsx")

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(row.to_dict() for row in rows)

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(HEADERS)
    for row in rows:
        data_sheet.append([row.to_dict()[header] for header in HEADERS])
    for cell in data_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    widths = [24, 14, 30, 14, 18, 12, 14, 50]
    for index, width in enumerate(widths, start=1):
        data_sheet.column_dimensions[get_column_letter(index)].width = width

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Total rows", len(rows)])
    summary.append(["Rows requiring review", sum(row.needs_review for row in rows)])
    summary.append(["Average OCR confidence", round(sum(row.confidence for row in rows) / len(rows), 4) if rows else 0])
    summary.append([])
    summary.append(["Blood group", "Count"])
    for blood_group, count in sorted(Counter(row.blood_group for row in rows if row.blood_group).items()):
        summary.append([blood_group, count])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 16
    workbook.save(xlsx_path)
    return csv_path, xlsx_path
