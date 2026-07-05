import csv

from openpyxl import load_workbook

from table_ocr.export import export_rows
from table_ocr.models import TableRow


def test_export_rows_writes_csv_and_xlsx(tmp_path):
    rows = [TableRow("one.jpg", "B-1", "Leena Desai", "A+", "8369476756", 0.91, False, "raw")]

    csv_path, xlsx_path = export_rows(rows, tmp_path / "donors")

    with csv_path.open(newline="", encoding="utf-8") as handle:
        records = list(csv.DictReader(handle))
    assert records[0]["donor_name"] == "Leena Desai"
    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["Data", "Summary"]
    assert workbook["Data"]["B1"].value == "bag"
    assert workbook["Data"]["C2"].value == "B-1"
